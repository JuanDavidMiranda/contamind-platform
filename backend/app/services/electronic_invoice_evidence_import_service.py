"""Importación auditable de estados electrónicos, sin conexión con DIAN."""

from __future__ import annotations

import csv
import hashlib
import unicodedata
from dataclasses import dataclass
from datetime import UTC, date, datetime
from io import BytesIO, StringIO
from uuid import UUID, uuid4
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models.accounting import InvoiceRecord
from app.models.electronic_invoicing import (
    ElectronicInvoiceEvidenceImportRecord,
    ElectronicInvoiceEvidenceImportRowRecord,
)
from app.shared.errors import AppError, app_error


_STATUS_GROUPS = {
    "accepted": frozenset({"accepted", "validated", "approved", "dian_accepted"}),
    "pending": frozenset({"draft", "issued", "sent", "submitted", "pending", "processing"}),
    "rejected": frozenset({"rejected", "error", "failed", "invalid"}),
}
_VALID_STATUSES = frozenset().union(*_STATUS_GROUPS.values())
_FIELD_ALIASES = {
    "invoice_number": ("invoice_number", "numero_factura", "numero", "factura"),
    "electronic_status": ("electronic_status", "estado_electronico", "estado"),
    "electronic_reference": (
        "electronic_reference",
        "referencia_electronica",
        "referencia",
        "cufe",
        "cude",
    ),
    "electronic_status_at": (
        "electronic_status_at",
        "fecha_respuesta",
        "fecha_estado",
    ),
}


@dataclass(frozen=True)
class ElectronicInvoiceEvidenceRejection:
    row_number: int
    message: str


@dataclass(frozen=True)
class ElectronicInvoiceEvidenceImportResult:
    import_id: UUID
    accepted_rows: int
    duplicate_rows: int
    rejections: tuple[ElectronicInvoiceEvidenceRejection, ...]


@dataclass(frozen=True)
class ElectronicInvoiceEvidenceImportItem:
    id: UUID
    file_format: str
    accepted_rows: int
    duplicate_rows: int
    rejected_rows: int
    created_at: datetime


@dataclass(frozen=True)
class ElectronicInvoiceEvidenceImportRowItem:
    row_number: int
    outcome: str
    reason: str | None
    created_at: datetime


@dataclass(frozen=True)
class ElectronicInvoiceExceptionItem:
    invoice_id: UUID
    invoice_number: str | None
    issue_date: date
    electronic_status: str | None
    electronic_status_at: datetime | None
    has_electronic_reference: bool
    issue_codes: tuple[str, ...]


@dataclass(frozen=True)
class ElectronicInvoiceExceptionsPage:
    total: int
    items: tuple[ElectronicInvoiceExceptionItem, ...]


class ElectronicInvoiceEvidenceImportService:
    """Actualiza evidencia importada y registra decisiones por fila de forma segura."""

    def __init__(self, db: Session, *, max_xlsx_uncompressed_bytes: int = 25_000_000) -> None:
        self._db = db
        self._max_xlsx_uncompressed_bytes = max_xlsx_uncompressed_bytes

    def import_content(
        self,
        company_id: UUID,
        content: bytes,
        *,
        uploaded_format: str,
        actor_user_id: int,
    ) -> ElectronicInvoiceEvidenceImportResult:
        if uploaded_format not in {"csv", "xlsx"}:
            raise app_error("VALIDATION_ERROR", message="El archivo debe ser CSV o XLSX.")
        rows = self._read_rows(content, uploaded_format)
        import_record = ElectronicInvoiceEvidenceImportRecord(
            id=str(uuid4()),
            company_id=str(company_id),
            file_format=uploaded_format,
            content_sha256=hashlib.sha256(content).hexdigest(),
            accepted_rows=0,
            duplicate_rows=0,
            rejected_rows=0,
            created_by_user_id=actor_user_id,
        )
        self._db.add(import_record)
        self._db.flush()

        accepted = duplicates = 0
        rejections: list[ElectronicInvoiceEvidenceRejection] = []
        seen_numbers: set[str] = set()
        for row_number, row in rows:
            invoice_id: str | None = None
            try:
                number = self._required(row, "invoice_number")
                number_key = number.casefold()
                if number_key in seen_numbers:
                    self._audit_row(
                        import_record,
                        row_number=row_number,
                        outcome="duplicate",
                        reason="La factura ya aparece en este archivo.",
                    )
                    duplicates += 1
                    continue
                seen_numbers.add(number_key)
                electronic_status = self._status(self._required(row, "electronic_status"))
                electronic_reference = self._optional(row, "electronic_reference", maximum=255)
                electronic_status_at = self._timestamp(row, "electronic_status_at")
                matches = list(
                    self._db.scalars(
                        select(InvoiceRecord).where(
                            InvoiceRecord.company_id == str(company_id),
                            InvoiceRecord.invoice_type == "sale",
                            func.lower(InvoiceRecord.number) == number_key,
                        )
                    )
                )
                if not matches:
                    raise ValueError("No existe una factura de venta con ese número en la empresa.")
                if len(matches) > 1:
                    raise ValueError("El número de factura es ambiguo y requiere revisión manual.")
                invoice = matches[0]
                invoice_id = invoice.id
                if self._is_same_evidence(
                    invoice,
                    electronic_status,
                    electronic_reference,
                    electronic_status_at,
                ):
                    self._audit_row(
                        import_record,
                        row_number=row_number,
                        outcome="duplicate",
                        invoice_id=invoice_id,
                        reason="La evidencia ya estaba registrada para esta factura.",
                    )
                    duplicates += 1
                    continue
                invoice.electronic_status = electronic_status
                invoice.electronic_reference = electronic_reference
                invoice.electronic_status_at = electronic_status_at
                invoice.updated_by_user_id = actor_user_id
                self._audit_row(
                    import_record,
                    row_number=row_number,
                    outcome="accepted",
                    invoice_id=invoice_id,
                )
                accepted += 1
            except (AppError, TypeError, ValueError) as exc:
                rejection = ElectronicInvoiceEvidenceRejection(
                    row_number=row_number,
                    message=(str(exc) or "Fila inválida.")[:500],
                )
                rejections.append(rejection)
                self._audit_row(
                    import_record,
                    row_number=row_number,
                    outcome="rejected",
                    invoice_id=invoice_id,
                    reason=rejection.message,
                )

        import_record.accepted_rows = accepted
        import_record.duplicate_rows = duplicates
        import_record.rejected_rows = len(rejections)
        self._db.commit()
        return ElectronicInvoiceEvidenceImportResult(
            import_id=UUID(import_record.id),
            accepted_rows=accepted,
            duplicate_rows=duplicates,
            rejections=tuple(rejections),
        )

    def list_imports(
        self,
        company_id: UUID,
        *,
        limit: int,
        offset: int,
    ) -> tuple[int, tuple[ElectronicInvoiceEvidenceImportItem, ...]]:
        filters = [ElectronicInvoiceEvidenceImportRecord.company_id == str(company_id)]
        total = int(
            self._db.scalar(select(func.count(ElectronicInvoiceEvidenceImportRecord.id)).where(*filters))
            or 0
        )
        records = self._db.scalars(
            select(ElectronicInvoiceEvidenceImportRecord)
            .where(*filters)
            .order_by(ElectronicInvoiceEvidenceImportRecord.created_at.desc())
            .offset(offset)
            .limit(limit)
        )
        return total, tuple(
            ElectronicInvoiceEvidenceImportItem(
                id=UUID(record.id),
                file_format=record.file_format,
                accepted_rows=record.accepted_rows,
                duplicate_rows=record.duplicate_rows,
                rejected_rows=record.rejected_rows,
                created_at=record.created_at,
            )
            for record in records
        )

    def list_import_rows(
        self,
        company_id: UUID,
        import_id: UUID,
        *,
        limit: int,
        offset: int,
    ) -> tuple[int, tuple[ElectronicInvoiceEvidenceImportRowItem, ...]]:
        filters = [
            ElectronicInvoiceEvidenceImportRowRecord.company_id == str(company_id),
            ElectronicInvoiceEvidenceImportRowRecord.import_id == str(import_id),
        ]
        exists = self._db.scalar(
            select(ElectronicInvoiceEvidenceImportRecord.id).where(
                ElectronicInvoiceEvidenceImportRecord.id == str(import_id),
                ElectronicInvoiceEvidenceImportRecord.company_id == str(company_id),
            )
        )
        if exists is None:
            raise app_error("NOT_FOUND", message="Importación de evidencia no encontrada para esta empresa.")
        total = int(
            self._db.scalar(select(func.count(ElectronicInvoiceEvidenceImportRowRecord.id)).where(*filters))
            or 0
        )
        records = self._db.scalars(
            select(ElectronicInvoiceEvidenceImportRowRecord)
            .where(*filters)
            .order_by(ElectronicInvoiceEvidenceImportRowRecord.row_number.asc())
            .offset(offset)
            .limit(limit)
        )
        return total, tuple(
            ElectronicInvoiceEvidenceImportRowItem(
                row_number=record.row_number,
                outcome=record.outcome,
                reason=record.reason,
                created_at=record.created_at,
            )
            for record in records
        )

    def exceptions(
        self,
        company_id: UUID,
        *,
        as_of: datetime | None = None,
        limit: int,
        offset: int,
    ) -> ElectronicInvoiceExceptionsPage:
        analysis_at = as_of or datetime.now(UTC).replace(tzinfo=None)
        records = tuple(
            self._db.scalars(
                select(InvoiceRecord)
                .where(
                    InvoiceRecord.company_id == str(company_id),
                    InvoiceRecord.invoice_type == "sale",
                )
                .order_by(InvoiceRecord.issue_date.desc(), InvoiceRecord.id.asc())
            )
        )
        items = tuple(
            item
            for record in records
            if (item := self._exception_item(record, analysis_at)) is not None
        )
        return ElectronicInvoiceExceptionsPage(total=len(items), items=items[offset:offset + limit])

    def _audit_row(
        self,
        import_record: ElectronicInvoiceEvidenceImportRecord,
        *,
        row_number: int,
        outcome: str,
        invoice_id: str | None = None,
        reason: str | None = None,
    ) -> None:
        import_record_row = ElectronicInvoiceEvidenceImportRowRecord(
            id=str(uuid4()),
            import_id=import_record.id,
            company_id=import_record.company_id,
            invoice_id=invoice_id,
            row_number=row_number,
            outcome=outcome,
            reason=reason,
        )
        self._db.add(import_record_row)

    def _read_rows(self, content: bytes, uploaded_format: str) -> tuple[tuple[int, dict[str, object]], ...]:
        if not content:
            raise app_error("VALIDATION_ERROR", message="El archivo está vacío.")
        if uploaded_format == "csv":
            try:
                text = content.decode("utf-8-sig")
            except UnicodeDecodeError as exc:
                raise app_error("VALIDATION_ERROR", message="El archivo CSV debe usar codificación UTF-8.") from exc
            if not text.strip():
                raise app_error("VALIDATION_ERROR", message="El archivo está vacío.")
            try:
                dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;")
            except csv.Error:
                dialect = csv.excel
            reader = csv.DictReader(StringIO(text), dialect=dialect)
            fields = self._resolve_fields(reader.fieldnames or [])
            return tuple(
                (row_number, self._canonical_row(row, fields))
                for row_number, row in enumerate(reader, start=2)
            )
        return self._read_xlsx(content)

    def _read_xlsx(self, content: bytes) -> tuple[tuple[int, dict[str, object]], ...]:
        try:
            with ZipFile(BytesIO(content)) as archive:
                if sum(item.file_size for item in archive.infolist()) > self._max_xlsx_uncompressed_bytes:
                    raise ValueError("El contenido XLSX excede el límite permitido.")
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
            try:
                rows = workbook.active.iter_rows(values_only=True)
                headers = next(rows, None)
                header_values = [str(value).strip() if value is not None else "" for value in headers or ()]
                fields = self._resolve_fields(header_values)
                return tuple(
                    (
                        row_number,
                        {
                            name: values[header_values.index(header)] if header_values.index(header) < len(values) else None
                            for name, header in fields.items()
                        },
                    )
                    for row_number, values in enumerate(rows, start=2)
                )
            finally:
                workbook.close()
        except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
            raise app_error("VALIDATION_ERROR", message="El archivo XLSX no es válido.") from exc

    @staticmethod
    def _resolve_fields(headers: list[str]) -> dict[str, str]:
        if not headers or not all(header.strip() for header in headers):
            raise app_error("VALIDATION_ERROR", message="Los encabezados no pueden estar vacíos.")
        normalized = [_normalize_header(header) for header in headers]
        if len(set(normalized)) != len(normalized):
            raise app_error("VALIDATION_ERROR", message="Los encabezados deben ser únicos.")
        by_name = {normalized_name: original for normalized_name, original in zip(normalized, headers, strict=True)}
        resolved: dict[str, str] = {}
        for field, aliases in _FIELD_ALIASES.items():
            match = next((by_name[alias] for alias in aliases if alias in by_name), None)
            if match is not None:
                resolved[field] = match
        required = {"invoice_number", "electronic_status"}
        missing = required - set(resolved)
        if missing:
            raise app_error(
                "VALIDATION_ERROR",
                message="El archivo debe incluir las columnas número de factura y estado electrónico.",
            )
        return resolved

    @staticmethod
    def _canonical_row(row: dict[str | None, str | None], fields: dict[str, str]) -> dict[str, object]:
        return {field: row.get(header) for field, header in fields.items()}

    @staticmethod
    def _required(row: dict[str, object], field: str) -> str:
        value = str(row.get(field) or "").strip()
        if not value:
            names = {
                "invoice_number": "El número de factura es obligatorio.",
                "electronic_status": "El estado electrónico es obligatorio.",
            }
            raise ValueError(names[field])
        return value

    @staticmethod
    def _optional(row: dict[str, object], field: str, *, maximum: int) -> str | None:
        value = str(row.get(field) or "").strip()
        if len(value) > maximum:
            raise ValueError("La referencia electrónica supera el máximo permitido de 255 caracteres.")
        return value or None

    @staticmethod
    def _status(value: str) -> str:
        normalized = _normalize_header(value)
        if normalized not in _VALID_STATUSES:
            raise ValueError("El estado electrónico no es reconocido.")
        return normalized

    @staticmethod
    def _timestamp(row: dict[str, object], field: str) -> datetime | None:
        value = row.get(field)
        if value is None or not str(value).strip():
            return None
        if isinstance(value, datetime):
            parsed = value
        else:
            try:
                parsed = datetime.fromisoformat(str(value).strip().replace("Z", "+00:00"))
            except ValueError as exc:
                raise ValueError("La fecha de respuesta debe usar formato ISO 8601.") from exc
        if parsed.tzinfo is not None:
            parsed = parsed.astimezone(UTC).replace(tzinfo=None)
        return parsed

    @staticmethod
    def _is_same_evidence(
        invoice: InvoiceRecord,
        status: str,
        reference: str | None,
        status_at: datetime | None,
    ) -> bool:
        return (
            invoice.electronic_status == status
            and invoice.electronic_reference == reference
            and invoice.electronic_status_at == status_at
        )

    @staticmethod
    def _exception_item(
        invoice: InvoiceRecord,
        analysis_at: datetime,
    ) -> ElectronicInvoiceExceptionItem | None:
        codes: list[str] = []
        status = _normalize_header(invoice.electronic_status or "")
        if status in _STATUS_GROUPS["rejected"]:
            codes.append("ELECTRONIC_STATUS_REJECTED")
        elif status in _STATUS_GROUPS["pending"]:
            codes.append("ELECTRONIC_STATUS_PENDING")
        elif status not in _STATUS_GROUPS["accepted"]:
            codes.append("ELECTRONIC_STATUS_MISSING")
        has_reference = bool(invoice.electronic_reference and invoice.electronic_reference.strip())
        if not has_reference:
            codes.append("ELECTRONIC_REFERENCE_MISSING")
        if not invoice.number or not invoice.number.strip():
            codes.append("INVOICE_NUMBER_MISSING")
        if invoice.recipient_party_id is None:
            codes.append("RECIPIENT_MISSING")
        expected_total = (invoice.subtotal + invoice.tax_total - invoice.withholding_total).quantize(
            invoice.total.as_tuple().exponent
        )
        if invoice.total != expected_total:
            codes.append("TOTAL_MISMATCH")
        if invoice.issue_date > analysis_at.date():
            codes.append("FUTURE_ISSUE_DATE")
        if not codes:
            return None
        return ElectronicInvoiceExceptionItem(
            invoice_id=UUID(invoice.id),
            invoice_number=invoice.number,
            issue_date=invoice.issue_date,
            electronic_status=invoice.electronic_status,
            electronic_status_at=invoice.electronic_status_at,
            has_electronic_reference=has_reference,
            issue_codes=tuple(codes),
        )


def _normalize_header(value: str) -> str:
    decomposed = unicodedata.normalize("NFKD", value.strip().casefold())
    without_marks = "".join(character for character in decomposed if not unicodedata.combining(character))
    return "_".join(without_marks.replace("-", " ").split())
