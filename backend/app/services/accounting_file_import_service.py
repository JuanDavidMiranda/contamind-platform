"""Importación tabular de entidades contables mediante perfiles explícitos."""

import csv
import hashlib
from collections import defaultdict
from datetime import date
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.data_sources.models import (
    AccountingImportResult,
    CompanyDataSource,
    ConnectionMode,
    DataCapability,
    DataSourceKind,
    FileFormat,
    ImportEntity,
    ImportProfile,
    ImportRejection,
)
from app.models.accounting import InvoiceRecord, ItemRecord, TaxRecord
from app.models.data_source import PartyRecord
from app.providers.canonical import Currency, InvoiceLine, InvoiceType, ItemType, JournalEntryLine
from app.services.manual_accounting_service import ManualAccountingService
from app.shared.errors import AppError, app_error


_CAPABILITY_BY_ENTITY = {
    ImportEntity.TAXES: DataCapability.TAXES,
    ImportEntity.ITEMS: DataCapability.ITEMS,
    ImportEntity.INVOICES: DataCapability.INVOICES,
    ImportEntity.PAYMENTS: DataCapability.PAYMENTS,
    ImportEntity.JOURNAL_ENTRIES: DataCapability.JOURNALS,
}


class AccountingFileImportService:
    """Convierte filas CSV/XLSX en llamadas al servicio canónico de contabilidad."""

    def __init__(self, db: Session, *, max_xlsx_uncompressed_bytes: int = 25_000_000) -> None:
        self._db = db
        self._manual = ManualAccountingService(db)
        self._max_xlsx_uncompressed_bytes = max_xlsx_uncompressed_bytes

    def import_content(
        self,
        source: CompanyDataSource,
        profile: ImportProfile,
        content: bytes,
        *,
        uploaded_format: FileFormat,
        actor_user_id: int,
    ) -> AccountingImportResult:
        self._validate_request(source, profile, uploaded_format)
        rows = self._read_rows(content, uploaded_format)
        content_hash = hashlib.sha256(content).hexdigest()
        if profile.entity is ImportEntity.TAXES:
            return self._single_rows(source, profile, rows, content_hash, actor_user_id, self._create_tax)
        if profile.entity is ImportEntity.ITEMS:
            return self._single_rows(source, profile, rows, content_hash, actor_user_id, self._create_item)
        if profile.entity is ImportEntity.PAYMENTS:
            return self._single_rows(source, profile, rows, content_hash, actor_user_id, self._create_payment)
        if profile.entity is ImportEntity.INVOICES:
            return self._grouped_invoices(source, profile, rows, content_hash, actor_user_id)
        if profile.entity is ImportEntity.JOURNAL_ENTRIES:
            return self._grouped_journals(source, profile, rows, content_hash, actor_user_id)
        raise app_error("CONFLICT", message="Esta entidad debe importarse por su ruta especializada.")

    def _single_rows(self, source, profile, rows, content_hash, actor_user_id, create) -> AccountingImportResult:
        accepted = 0
        rejections: list[ImportRejection] = []
        for row_number, row in rows:
            try:
                create(source, profile, row, actor_user_id, self._idempotency_key(content_hash, profile.entity, str(row_number)))
                accepted += 1
            except (AppError, ValueError, TypeError, InvalidOperation) as exc:
                rejections.append(self._rejection(row_number, exc))
        return AccountingImportResult(
            entity=profile.entity,
            accepted_rows=accepted,
            rejections=tuple(rejections),
        )

    def _grouped_invoices(self, source, profile, rows, content_hash, actor_user_id) -> AccountingImportResult:
        groups, rejections = self._groups(profile, rows, "number")
        accepted = 0
        for number, group in groups.items():
            try:
                first = group[0][1]
                lines = tuple(self._invoice_line(profile, row, source.company_id) for _, row in group)
                self._manual.create_invoice(
                    source.id,
                    invoice_type=InvoiceType(self._required(profile, first, "invoice_type")),
                    issue_date=self._date(profile, first, "issue_date", required=True),
                    issuer_party_id=self._party_id(profile, first, source.company_id, "issuer"),
                    recipient_party_id=self._party_id(profile, first, source.company_id, "recipient"),
                    lines=lines,
                    currency=self._currency(profile, first),
                    tax_total=self._decimal(profile, first, "tax_total", default=Decimal("0")),
                    withholding_total=self._decimal(profile, first, "withholding_total", default=Decimal("0")),
                    number=number,
                    status=self._value(profile, first, "status"),
                    actor_user_id=actor_user_id,
                    idempotency_key=self._idempotency_key(content_hash, profile.entity, number),
                )
                accepted += len(group)
            except (AppError, ValueError, TypeError, InvalidOperation) as exc:
                rejections.extend(self._rejection(row_number, exc) for row_number, _ in group)
        return AccountingImportResult(
            entity=profile.entity,
            accepted_rows=accepted,
            rejections=tuple(rejections),
        )

    def _grouped_journals(self, source, profile, rows, content_hash, actor_user_id) -> AccountingImportResult:
        groups, rejections = self._groups(profile, rows, "source_reference")
        accepted = 0
        for reference, group in groups.items():
            try:
                first = group[0][1]
                lines = tuple(self._journal_line(profile, row, source.company_id) for _, row in group)
                self._manual.create_journal_entry(
                    source.id,
                    entry_date=self._date(profile, first, "entry_date", required=True),
                    description=self._required(profile, first, "description"),
                    lines=lines,
                    source_reference=reference,
                    actor_user_id=actor_user_id,
                    idempotency_key=self._idempotency_key(content_hash, profile.entity, reference),
                )
                accepted += len(group)
            except (AppError, ValueError, TypeError, InvalidOperation) as exc:
                rejections.extend(self._rejection(row_number, exc) for row_number, _ in group)
        return AccountingImportResult(
            entity=profile.entity,
            accepted_rows=accepted,
            rejections=tuple(rejections),
        )

    def _create_tax(self, source, profile, row, actor_user_id, idempotency_key) -> None:
        self._manual.create_tax(
            source.id,
            code=self._required(profile, row, "code"),
            name=self._required(profile, row, "name"),
            rate=self._decimal(profile, row, "rate", required=True),
            actor_user_id=actor_user_id,
            idempotency_key=idempotency_key,
        )

    def _create_item(self, source, profile, row, actor_user_id, idempotency_key) -> None:
        self._manual.create_item(
            source.id,
            code=self._required(profile, row, "code"),
            name=self._required(profile, row, "name"),
            item_type=ItemType(self._required(profile, row, "item_type")),
            unit=self._value(profile, row, "unit"),
            unit_price=self._decimal(profile, row, "unit_price", required=True),
            tax_ids=self._tax_ids(profile, row, source.company_id, "tax_codes"),
            ledger_account=self._value(profile, row, "ledger_account"),
            actor_user_id=actor_user_id,
            idempotency_key=idempotency_key,
        )

    def _create_payment(self, source, profile, row, actor_user_id, idempotency_key) -> None:
        self._manual.create_payment(
            source.id,
            payment_date=self._date(profile, row, "payment_date", required=True),
            amount=self._decimal(profile, row, "amount", required=True),
            currency=self._currency(profile, row),
            invoice_id=self._invoice_id(profile, row, source.company_id),
            payment_method=self._value(profile, row, "payment_method"),
            actor_user_id=actor_user_id,
            idempotency_key=idempotency_key,
        )

    def _invoice_line(self, profile, row, company_id) -> InvoiceLine:
        return InvoiceLine(
            item_id=self._item_id(profile, row, company_id),
            description=self._required(profile, row, "description"),
            quantity=self._decimal(profile, row, "quantity", required=True),
            unit_price=self._decimal(profile, row, "unit_price", required=True),
            tax_ids=self._tax_ids(profile, row, company_id, "tax_codes"),
            withholding_ids=self._tax_ids(profile, row, company_id, "withholding_codes"),
        )

    def _journal_line(self, profile, row, company_id) -> JournalEntryLine:
        return JournalEntryLine(
            account_code=self._required(profile, row, "account_code"),
            debit=self._decimal(profile, row, "debit", default=Decimal("0")),
            credit=self._decimal(profile, row, "credit", default=Decimal("0")),
            party_id=self._party_id(profile, row, company_id, "party"),
            cost_center=self._value(profile, row, "cost_center"),
        )

    def _groups(self, profile, rows, field):
        groups: dict[str, list[tuple[int, dict[str, str | None]]]] = defaultdict(list)
        rejections: list[ImportRejection] = []
        for row_number, row in rows:
            try:
                groups[self._required(profile, row, field)].append((row_number, row))
            except ValueError as exc:
                rejections.append(self._rejection(row_number, exc))
        return groups, rejections

    def _party_id(self, profile, row, company_id, prefix):
        explicit = self._value(profile, row, f"{prefix}_party_id")
        if explicit:
            return self._uuid(explicit, f"{prefix}_party_id")
        document = self._value(profile, row, f"{prefix}_document_number")
        if not document:
            return None
        record = self._db.scalar(
            select(PartyRecord).where(
                PartyRecord.company_id == str(company_id), PartyRecord.document_number == document
            )
        )
        if record is None:
            raise ValueError(f"No existe un tercero con documento '{document}'.")
        return self._uuid(record.id, f"{prefix}_party_id")

    def _item_id(self, profile, row, company_id):
        explicit = self._value(profile, row, "item_id")
        if explicit:
            return self._uuid(explicit, "item_id")
        code = self._value(profile, row, "item_code")
        if not code:
            return None
        record = self._db.scalar(
            select(ItemRecord).where(ItemRecord.company_id == str(company_id), ItemRecord.code == code)
        )
        if record is None:
            raise ValueError(f"No existe un ítem con código '{code}'.")
        return self._uuid(record.id, "item_id")

    def _invoice_id(self, profile, row, company_id):
        explicit = self._value(profile, row, "invoice_id")
        if explicit:
            return self._uuid(explicit, "invoice_id")
        number = self._value(profile, row, "invoice_number")
        if not number:
            return None
        record = self._db.scalar(
            select(InvoiceRecord).where(
                InvoiceRecord.company_id == str(company_id), InvoiceRecord.number == number
            )
        )
        if record is None:
            raise ValueError(f"No existe una factura con número '{number}'.")
        return self._uuid(record.id, "invoice_id")

    def _tax_ids(self, profile, row, company_id, field):
        explicit = self._value(profile, row, field.replace("_codes", "_ids"))
        if explicit:
            return tuple(self._uuid(value.strip(), field) for value in explicit.split(",") if value.strip())
        codes = self._value(profile, row, field)
        if not codes:
            return ()
        values = [value.strip() for value in codes.split(",") if value.strip()]
        records = list(
            self._db.scalars(
                select(TaxRecord).where(
                    TaxRecord.company_id == str(company_id), TaxRecord.code.in_(values)
                )
            )
        )
        by_code = {record.code: record for record in records}
        missing = [code for code in values if code not in by_code]
        if missing:
            raise ValueError(f"No existe el impuesto con código '{missing[0]}'.")
        return tuple(self._uuid(by_code[code].id, field) for code in values)

    def _currency(self, profile, row) -> Currency:
        return Currency(
            code=self._value(profile, row, "currency_code") or "COP",
            exchange_rate=self._decimal(profile, row, "exchange_rate", default=Decimal("1")),
            as_of=self._date(profile, row, "currency_as_of"),
        )

    @staticmethod
    def _validate_request(source, profile, uploaded_format):
        if source.kind is not DataSourceKind.FILE_IMPORT or source.mode is not ConnectionMode.FILE_UPLOAD:
            raise app_error("CONFLICT", message="La fuente no está configurada para importar archivos.")
        if profile.data_source_id != source.id or profile.file_format is not uploaded_format:
            raise app_error("CONFLICT", message="El perfil no corresponde a la fuente o formato cargado.")
        expected_connector = "csv_import" if uploaded_format is FileFormat.CSV else "xlsx_import"
        if source.connector_id != expected_connector:
            raise app_error("CONFLICT", message="La fuente no está configurada para este formato de archivo.")
        if profile.entity not in _CAPABILITY_BY_ENTITY:
            raise app_error("CONFLICT", message="El perfil debe apuntar a una entidad contable importable.")
        if _CAPABILITY_BY_ENTITY[profile.entity] not in source.capabilities:
            raise app_error("CONFLICT", message="La fuente no tiene habilitada esta capacidad.")

    def _read_rows(self, content: bytes, file_format: FileFormat):
        if file_format is FileFormat.CSV:
            try:
                reader = csv.DictReader(StringIO(content.decode("utf-8-sig")))
            except UnicodeDecodeError as exc:
                raise app_error("VALIDATION_ERROR", message="El archivo CSV debe usar codificación UTF-8.") from exc
            self._validate_headers(reader.fieldnames)
            return [(index, row) for index, row in enumerate(reader, start=2)]
        return self._read_xlsx(content)

    def _read_xlsx(self, content: bytes):
        try:
            with ZipFile(BytesIO(content)) as archive:
                if sum(item.file_size for item in archive.infolist()) > self._max_xlsx_uncompressed_bytes:
                    raise ValueError("El contenido XLSX excede el límite permitido.")
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
            try:
                rows = workbook.active.iter_rows(values_only=True)
                headers = next(rows, None)
                normalized = [str(value).strip() if value is not None else "" for value in headers or ()]
                self._validate_headers(normalized)
                return [
                    (
                        row_number,
                        {
                            normalized[index]: str(value).strip() if value is not None else None
                            for index, value in enumerate(values)
                            if index < len(normalized)
                        },
                    )
                    for row_number, values in enumerate(rows, start=2)
                ]
            finally:
                workbook.close()
        except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
            raise app_error("VALIDATION_ERROR", message="El archivo XLSX no es válido.") from exc

    @staticmethod
    def _validate_headers(headers):
        if not headers or not all(headers) or len(set(headers)) != len(headers):
            raise app_error("VALIDATION_ERROR", message="Los encabezados deben ser únicos y no vacíos.")

    @staticmethod
    def _idempotency_key(content_hash, entity, discriminator):
        return hashlib.sha256(f"{content_hash}:{entity.value}:{discriminator}".encode()).hexdigest()

    @staticmethod
    def _rejection(row_number, exc):
        return ImportRejection(row_number=row_number, message=str(exc)[:500] or "Fila inválida.")

    @staticmethod
    def _uuid(value, field):
        try:
            from uuid import UUID

            return UUID(value)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"'{field}' debe ser un UUID válido.") from exc

    @staticmethod
    def _value(profile, row, field):
        column = profile.column_mapping.get(field)
        if column is None:
            return None
        value = (row.get(column) or "").strip()
        return value or None

    def _required(self, profile, row, field):
        value = self._value(profile, row, field)
        if value is None:
            column = profile.column_mapping.get(field)
            if column is None:
                raise ValueError(f"Falta el mapeo obligatorio para '{field}'.")
            raise ValueError(f"La columna '{column}' es obligatoria.")
        return value

    def _decimal(self, profile, row, field, *, required=False, default=None):
        value = self._required(profile, row, field) if required else self._value(profile, row, field)
        if value is None:
            if default is not None:
                return default
            return None
        try:
            return Decimal(value.replace(".", "").replace(",", ".")) if "," in value else Decimal(value)
        except InvalidOperation as exc:
            raise ValueError(f"'{field}' debe ser un número válido.") from exc

    def _date(self, profile, row, field, *, required=False):
        value = self._required(profile, row, field) if required else self._value(profile, row, field)
        if value is None:
            return None
        try:
            return date.fromisoformat(value)
        except ValueError as exc:
            raise ValueError(f"'{field}' debe usar formato AAAA-MM-DD.") from exc
