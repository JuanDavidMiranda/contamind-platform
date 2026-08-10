"""Ingreso seguro de terceros desde libros XLSX mediante perfiles de mapeo."""

from io import BytesIO
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.data_sources.csv_import import CsvPartyImportSource
from app.data_sources.models import (
    CompanyDataSource,
    ConnectionMode,
    DataSourceContext,
    FileFormat,
    ImportProfile,
    PartyImportResult,
)
from app.shared.errors import app_error


class XlsxPartyImportSource(CsvPartyImportSource):
    """Lee solo valores del libro; fórmulas, macros y estilos no se ejecutan ni persisten."""

    connector_id = "xlsx_import"
    supported_modes = frozenset({ConnectionMode.FILE_UPLOAD})

    def __init__(self, max_uncompressed_bytes: int = 25_000_000) -> None:
        super().__init__()
        self._max_uncompressed_bytes = max_uncompressed_bytes

    async def import_parties(
        self,
        context: DataSourceContext,
        source: CompanyDataSource,
        profile: ImportProfile,
        content: bytes,
    ) -> PartyImportResult:
        self._validate_request(context, source, profile, FileFormat.XLSX)
        try:
            with ZipFile(BytesIO(content)) as archive:
                if sum(item.file_size for item in archive.infolist()) > self._max_uncompressed_bytes:
                    raise ValueError("El contenido XLSX excede el límite permitido.")
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
            try:
                rows = workbook.active.iter_rows(values_only=True)
                header_values = next(rows, None)
                if header_values is None:
                    raise ValueError("El libro no contiene encabezados.")
                headers = [str(value).strip() if value is not None else "" for value in header_values]
                if not any(headers) or len(set(headers)) != len(headers):
                    raise ValueError("Los encabezados del libro deben ser únicos y no vacíos.")
                mapped_rows = (
                    (
                        row_number,
                        {
                            headers[index]: str(value).strip() if value is not None else None
                            for index, value in enumerate(values)
                        },
                    )
                    for row_number, values in enumerate(rows, start=2)
                )
                return self._process_rows(context, source, profile, mapped_rows)
            finally:
                workbook.close()
        except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
            raise app_error("VALIDATION_ERROR", message="El archivo XLSX no es válido.") from exc
